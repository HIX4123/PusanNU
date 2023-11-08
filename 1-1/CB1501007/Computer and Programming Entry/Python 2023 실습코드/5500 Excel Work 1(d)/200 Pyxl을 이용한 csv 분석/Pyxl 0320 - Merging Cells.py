#-------------------------------------------------------------------------------
# Author:      Zoh
# Created:     2019-07-06
#-------------------------------------------------------------------------------
# To get those cells that actually contain data, we can use dimensions.
# dimensions.py

from openpyxl import Workbook
from openpyxl.styles import Alignment

book = Workbook()
sheet = book.active

sheet.merge_cells('A1:B2')

cell = sheet.cell(row=1, column=1)
cell.value = 'Sunny day'
cell.alignment = Alignment(horizontal='center', vertical='center')

book.save('merging.xlsx')