#-------------------------------------------------------------------------------
# Author:      Zoh
# Created:     2019-07-06
#-------------------------------------------------------------------------------
# To get those cells that actually contain data, we can use dimensions.
# dimensions.py
#!/usr/bin/python3

import openpyxl

book = openpyxl.load_workbook('sheets.xlsx')

book.create_sheet("April")

print(book.sheetnames)

sheet1 = book.get_sheet_by_name("Tomato")
book.remove_sheet(sheet1)

print(book.sheetnames)

book.create_sheet("January", 0)
print(book.sheetnames)

book.save('sheets2.xlsx')
