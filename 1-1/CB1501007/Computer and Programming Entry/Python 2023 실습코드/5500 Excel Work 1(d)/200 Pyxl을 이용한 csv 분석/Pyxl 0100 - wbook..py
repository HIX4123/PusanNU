#-------------------------------------------------------------------------------
# Author:      Zoh
# Created:     2019-07-06
#-------------------------------------------------------------------------------


from openpyxl import Workbook
wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# or
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position

ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]

print(wb.sheetnames)

for sheet in wb:
    print("Title=", sheet.title)

